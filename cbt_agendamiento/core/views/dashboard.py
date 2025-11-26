from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.db.models import Count, Q
from django.contrib.gis.geos import Point
from datetime import date, datetime
from django.db import transaction
from django.http import HttpResponse

# Excel
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from ..models import Turno, Establecimiento, Notificacion
from ..utils import enviar_correo_html

def es_staff(user): return user.is_staff

@login_required
@user_passes_test(es_staff)
def dashboard_staff(request):
    stats = Turno.objects.aggregate(
        pendientes=Count('id', filter=Q(estado='PENDIENTE')),
        confirmados=Count('id', filter=Q(estado='CONFIRMADO')),
        rechazados=Count('id', filter=Q(estado='RECHAZADO')),
    )

    pendientes = Turno.objects.filter(estado='PENDIENTE')\
        .select_related('establecimiento__propietario', 'agenda')\
        .order_by('agenda__fecha', 'bloque')

    proximos = Turno.objects.filter(
        estado='CONFIRMADO', 
        agenda__fecha__gte=date.today()
    ).select_related('establecimiento__propietario', 'agenda').order_by('agenda__fecha')[:10]

    context = {
        'kpi_locales': Establecimiento.objects.count(),
        'kpi_hoy': Turno.objects.filter(agenda__fecha=date.today()).count(),
        'stats': stats, 
        'lista_pendientes': pendientes,
        'lista_proximos': proximos,
        'hoy': date.today(),
        'mes_actual': date.today().month,
        'anio_actual': date.today().year
    }
    return render(request, 'staff/dashboard.html', context)

@login_required
@user_passes_test(es_staff)
def gestionar_turno(request, turno_id, accion):
    with transaction.atomic():
        try:
            turno = Turno.objects.select_for_update().get(id=turno_id)
        except Turno.DoesNotExist:
            return redirect('dashboard_staff')

        if turno.estado != 'PENDIENTE':
            messages.warning(request, "Turno ya procesado.")
            return redirect('dashboard_staff')

        email = turno.establecimiento.propietario.email 
        propietario = turno.establecimiento.propietario
        
        datos_email = {
            'nombre': propietario.first_name,
            'local': turno.establecimiento.nombre_comercial,
            'fecha': turno.agenda.fecha.strftime("%d/%m/%Y"),
            'jornada': turno.get_bloque_display(),
        }
        
        if accion == 'confirmar':
            turno.estado = 'CONFIRMADO'
            turno.inspector = request.user 
            
            Notificacion.objects.create(
                usuario=propietario, titulo="¡Turno Aprobado! ✅",
                mensaje=f"Inspección confirmada para el {turno.agenda.fecha}.", tipo="SUCCESS", link="/portal/"
            )
            
            subject = "Inspección Confirmada ✅"
            datos_email['mensaje_principal'] = "Solicitud APROBADA."
            datos_email['estado'] = "CONFIRMADO"
            datos_email['color_estado'] = "#198754"
            datos_email['instrucciones'] = "Esté presente en el local."
        
        elif accion == 'rechazar':
            turno.estado = 'RECHAZADO'
            
            Notificacion.objects.create(
                usuario=propietario, titulo="Solicitud Rechazada ⚠️",
                mensaje="No pudimos procesar su turno.", tipo="WARNING", link="/portal/"
            )

            subject = "Actualización de Solicitud ⚠️"
            datos_email['mensaje_principal'] = "Solicitud no procesada."
            datos_email['estado'] = "RECHAZADO"
            datos_email['color_estado'] = "#dc3545"
            datos_email['instrucciones'] = "Seleccione nueva fecha."
        
        turno.save()
    
    if email: enviar_correo_html(email, subject, datos_email)
    return redirect('dashboard_staff')

@login_required
@user_passes_test(es_staff)
def finalizar_turno(request, turno_id):
    if request.method == 'POST':
        turno = get_object_or_404(Turno, id=turno_id)
        num = request.POST.get('numero_formulario')
        if num:
            turno.numero_formulario = num
            turno.estado = 'TERMINADO'
            turno.save()
            messages.success(request, f"Inspección cerrada. Formulario N° {num}.")
        else:
            messages.error(request, "Ingrese el número.")
    return redirect('dashboard_staff')

@login_required
@user_passes_test(es_staff)
def hoja_ruta(request):
    origen = Point(-77.7071697, 0.8234943, srid=4326) 
    bloque = request.GET.get('bloque', 'MANANA')
    zona = request.GET.get('zona', 'SUR')
    
    try: fecha = datetime.strptime(request.GET.get('fecha', ''), '%Y-%m-%d').date()
    except: fecha = date.today()

    filtro = 'GONZALEZ_SUAREZ' if zona == 'NORTE' else 'TULCAN_CENTRO'

    pendientes = list(Turno.objects.filter(
        agenda__fecha=fecha, estado='CONFIRMADO',
        establecimiento__parroquia=filtro, bloque=bloque
    ).select_related('establecimiento'))

    ruta_optimizada = []
    punto_actual = origen
    
    while pendientes:
        validos = [p for p in pendientes if p.establecimiento.ubicacion]
        if not validos: break 

        siguiente = min(validos, key=lambda t: t.establecimiento.ubicacion.distance(punto_actual))
        ruta_optimizada.append(siguiente)
        
        if siguiente in pendientes: pendientes.remove(siguiente) 
        punto_actual = siguiente.establecimiento.ubicacion

    return render(request, 'staff/hoja_ruta.html', {
        'ruta': ruta_optimizada, 'hoy': fecha, 'bloque_actual': bloque, 'zona_actual': zona
    })

@login_required
@user_passes_test(es_staff)
def generar_informe_mensual(request):
    try:
        mes = int(request.GET.get('mes', date.today().month))
        anio = int(request.GET.get('anio', date.today().year))
    except:
        mes = date.today().month
        anio = date.today().year

    turnos = Turno.objects.filter(
        estado='FINALIZADO',
        agenda__fecha__year=anio,
        agenda__fecha__month=mes
    ).select_related('establecimiento__propietario', 'agenda', 'inspector').order_by('agenda__fecha')

    nombres_meses = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]

    return render(request, 'staff/informe_mensual.html', {
        'turnos': turnos, 'mes': mes, 'anio': anio, 'nombre_mes': nombres_meses[mes], 'total': turnos.count()
    })

@login_required
@user_passes_test(es_staff)
def exportar_excel_mensual(request):
    try:
        mes = int(request.GET.get('mes', date.today().month))
        anio = int(request.GET.get('anio', date.today().year))
    except ValueError:
        mes = date.today().month
        anio = date.today().year

    turnos = Turno.objects.filter(
        estado='FINALIZADO',
        agenda__fecha__year=anio,
        agenda__fecha__month=mes
    ).select_related('establecimiento__propietario', 'agenda', 'inspector').order_by('agenda__fecha')

    # Crear Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Inspecciones {mes}-{anio}"

    font_title = Font(name='Calibri', size=14, bold=True, color='000000')
    font_header = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    fill_header = PatternFill(start_color='B02A37', end_color='B02A37', fill_type='solid')
    alignment_center = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.merge_cells('A1:H1')
    ws['A1'] = f"REPORTE MENSUAL DE INSPECCIONES - {mes}/{anio}"
    ws['A1'].font = font_title
    ws['A1'].alignment = alignment_center

    headers = ['N°', 'Fecha', 'N° Formulario', 'Nombre Comercial', 'Dirección', 'RUC', 'Inspector', 'Observaciones']
    ws.append([])
    ws.append(headers)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = alignment_center
        cell.border = thin_border

    for idx, turno in enumerate(turnos, 1):
        inspector_name = f"{turno.inspector.first_name} {turno.inspector.last_name}" if turno.inspector else "--"
        ruc = getattr(turno.establecimiento.propietario.perfil, 'ruc', 'N/A')
        
        row = [
            idx,
            turno.agenda.fecha,
            turno.numero_formulario or "S/N",
            turno.establecimiento.nombre_comercial,
            turno.establecimiento.direccion,
            ruc,
            inspector_name,
            turno.observaciones or ""
        ]
        ws.append(row)
        
        current_row = idx + 3
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.border = thin_border
            if col_num in [1, 2, 3, 6]: cell.alignment = alignment_center

    column_widths = [5, 12, 15, 35, 40, 15, 25, 30]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Reporte_CBT_{mes}_{anio}.xlsx'
    wb.save(response)
    return response