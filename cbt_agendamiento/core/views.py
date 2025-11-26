from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.contrib.auth.models import User
from django.contrib.gis.geos import Point
from datetime import date, datetime, timedelta
from django.db import IntegrityError, transaction
from django.db.models import Count, Q, ProtectedError
from django.http import JsonResponse, HttpResponse
import json
from django.core.serializers.json import DjangoJSONEncoder
from django.core.paginator import Paginator

# Imports Excel
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .utils import enviar_correo_html
from .forms import (
    AltaContribuyenteForm, TipoEstablecimientoForm, EdicionAgendaForm, 
    EditarUsuarioForm, NuevoInspectorForm, ConfiguracionGlobalForm, 
    RegistroEmailForm, MiPerfilForm,TasaPagoForm,RequisitoLegalForm
)
from .models import (
    Turno, Establecimiento, AgendaDiaria, TipoEstablecimiento, 
    OPCIONES_PARROQUIA, ConfiguracionSistema, PerfilUsuario, Notificacion, TasaPago,
    RequisitoLegal
)

def es_staff(user): return user.is_staff
def es_superuser(user): return user.is_superuser

# ==============================================================================
#                            PANEL DE CONTROL (STAFF)
# ==============================================================================

# 1. DASHBOARD PRINCIPAL (Solo Resumen)
@login_required
@user_passes_test(es_staff)
def dashboard_staff(request):
    # 1. KPIs (CORREGIDO: Se agregó 'rechazados' que faltaba)
    stats = Turno.objects.aggregate(
        pendientes=Count('id', filter=Q(estado='PENDIENTE')),
        confirmados=Count('id', filter=Q(estado='CONFIRMADO')),
        en_tramite=Count('id', filter=Q(estado='EJECUTADA')), # Estado intermedio
        rechazados=Count('id', filter=Q(estado='RECHAZADO')), # <--- ¡AQUÍ ESTABA EL ERROR!
    )

    # 2. LISTA DE PENDIENTES
    pendientes_list = Turno.objects.filter(estado='PENDIENTE')\
        .select_related('establecimiento__propietario', 'agenda')\
        .order_by('agenda__fecha', 'bloque', 'id')

    paginator = Paginator(pendientes_list, 6) 
    page_number = request.GET.get('page')
    pendientes_page = paginator.get_page(page_number)

    # 3. PRÓXIMOS (Agenda Lista - Incluye Confirmados Futuros/Hoy y Ejecutados)
    lista_cierre = Turno.objects.filter(
        Q(estado='CONFIRMADO', agenda__fecha__gte=date.today()) | 
        Q(estado='EJECUTADA')
    ).select_related('establecimiento__propietario', 'agenda').order_by('agenda__fecha')

    # 4. DATOS PARA VISUALIZACIÓN (JSON)
    agendas_futuras = AgendaDiaria.objects.filter(fecha__gte=date.today())
    eventos_calendario = []
    
    for ag in agendas_futuras:
        ocupacion = ag.turnos.filter(estado='CONFIRMADO').count()
        total = ag.capacidad_manana + ag.capacidad_tarde
        titulo = f"{ocupacion}/{total}"
        color = '#ef4444' if ocupacion >= total else '#10b981'
        
        eventos_calendario.append({
            'title': titulo,
            'start': ag.fecha.strftime("%Y-%m-%d"),
            'color': color,
            'url': f"/panel-operativo/agenda/editar/{ag.id}/"
        })
        
    puntos_mapa = []
    for turno in lista_cierre:
        if turno.establecimiento.ubicacion:
            puntos_mapa.append({
                'lat': turno.establecimiento.ubicacion.y,
                'lng': turno.establecimiento.ubicacion.x,
                'nombre': turno.establecimiento.nombre_comercial,
                'tipo': 'CONFIRMADO' if turno.estado == 'CONFIRMADO' else 'EJECUTADA'
            })

    context = {
        'kpi_locales': Establecimiento.objects.count(),
        'kpi_hoy': Turno.objects.filter(agenda__fecha=date.today()).count(),
        'stats': stats, 
        'lista_pendientes': pendientes_page,
        'lista_proximos': lista_cierre,
        'hoy': date.today(),
        'calendar_events_json': json.dumps(eventos_calendario, cls=DjangoJSONEncoder),
        'map_points_json': json.dumps(puntos_mapa, cls=DjangoJSONEncoder),
    }
    return render(request, 'staff/dashboard.html', context)

@login_required
@user_passes_test(es_staff)
def marcar_ejecutada(request, turno_id):
    if request.method == 'POST':
        turno = get_object_or_404(Turno, id=turno_id)
        
        # Solo si es hoy o antes (no se puede ejecutar futuro)
        if turno.agenda.fecha > date.today():
            messages.error(request, "No puede ejecutar inspecciones futuras.")
            return redirect('dashboard_staff')

        turno.estado = 'EJECUTADA'
        turno.save()
        
        # Notificar al ciudadano que la visita ocurrió
        Notificacion.objects.create(
            usuario=turno.establecimiento.propietario,
            titulo="Visita Realizada 🚒",
            mensaje="El inspector ha registrado la visita. Procesando informe final.",
            tipo="INFO", link="/portal/"
        )
        messages.success(request, f"Visita a {turno.establecimiento.nombre_comercial} registrada. Pendiente N° Formulario.")
            
    return redirect('dashboard_staff')

# 2. SOLICITUDES ENTRANTES
@login_required
@user_passes_test(es_staff)
def solicitudes_pendientes(request):
    # Orden: Fecha -> Bloque -> ID (Orden de llegada)
    pendientes_list = Turno.objects.filter(estado='PENDIENTE')\
        .select_related('establecimiento__propietario', 'agenda')\
        .order_by('agenda__fecha', 'bloque', 'id')
        
    # Paginación: 9 tarjetas por página (Grid de 3x3)
    paginator = Paginator(pendientes_list, 9)
    page_number = request.GET.get('page')
    pendientes_page = paginator.get_page(page_number)
        
    return render(request, 'staff/solicitudes.html', {'pendientes': pendientes_page})

# 3. GESTIÓN DE INSPECCIONES (CONFIRMADAS Y CANCELACIÓN)
@login_required
@user_passes_test(es_staff)
def gestion_inspecciones(request):
    # PESTAÑA 1: INSPECCIONES PROGRAMADAS (CONFIRMADAS FUTURAS O DE HOY)
    # Estas no se paginan porque son la carga de trabajo activa (pocas)
    programadas = Turno.objects.filter(
        estado='CONFIRMADO',
        agenda__fecha__gte=date.today()
    ).select_related('establecimiento', 'agenda').order_by('agenda__fecha')

    # PESTAÑA 2: HISTORIAL (FILTROS + PAGINACIÓN)
    q_search = request.GET.get('q', '')
    q_estado = request.GET.get('estado', '')
    
    # Excluir pendientes y confirmados (solo lo ya procesado)
    historial_list = Turno.objects.exclude(estado__in=['PENDIENTE', 'CONFIRMADO'])

    if q_search:
        historial_list = historial_list.filter(
            Q(establecimiento__nombre_comercial__icontains=q_search) |
            Q(establecimiento__propietario__perfil__ruc__icontains=q_search)
        )
    
    if q_estado:
        historial_list = historial_list.filter(estado=q_estado)

    # Ordenar por fecha descendente (lo más nuevo primero)
    historial_list = historial_list.select_related('establecimiento', 'agenda').order_by('-agenda__fecha')
    
    # Paginación: 20 registros por página
    paginator = Paginator(historial_list, 20)
    page_number = request.GET.get('page')
    historial_page = paginator.get_page(page_number)

    return render(request, 'staff/gestion_inspecciones.html', {
        'programadas': programadas,
        'historial': historial_page, # Pasamos la página, no la lista completa
        'filtros': {'q': q_search, 'estado': q_estado}
    })

@login_required
@user_passes_test(es_staff)
def cancelar_inspeccion_staff(request):
    if request.method == 'POST':
        turno_id = request.POST.get('turno_id')
        motivo = request.POST.get('motivo')
        turno = get_object_or_404(Turno, id=turno_id)
        
        if turno.estado != 'CONFIRMADO':
            messages.error(request, "Solo se pueden cancelar turnos confirmados.")
            return redirect('gestion_inspecciones')

        turno.estado = 'CANCELADO'
        turno.motivo_cancelacion = motivo
        turno.save()
        
        # Notificar
        Notificacion.objects.create(
            usuario=turno.establecimiento.propietario,
            titulo="Inspección Cancelada ❌",
            mensaje=f"Su turno ha sido cancelado. Motivo: {motivo}",
            tipo="ERROR", link="/portal/"
        )
        messages.success(request, "Inspección cancelada correctamente.")
        
    return redirect('gestion_inspecciones')

# 4. CIERRE DE INSPECCIONES
@login_required
@user_passes_test(es_staff)
def cierre_inspecciones(request):
    # 1. Capturar parámetros de filtro
    q = request.GET.get('q', '')
    estado_filter = request.GET.get('estado', 'todos') # 'todos', 'ruta', 'informe'
    
    # 2. Consulta Base: Solo los activos para cierre (CONFIRMADO o EJECUTADA)
    turnos_list = Turno.objects.filter(
        estado__in=['CONFIRMADO', 'EJECUTADA']
    ).select_related(
        'establecimiento__propietario__perfil', 
        'agenda', 
        'inspector'
    ).order_by('agenda__fecha', 'bloque') # Prioridad por fecha
    
    # 3. Aplicar Filtros
    if q:
        turnos_list = turnos_list.filter(
            Q(establecimiento__nombre_comercial__icontains=q) |
            Q(establecimiento__propietario__perfil__ruc__icontains=q) |
            Q(establecimiento__direccion__icontains=q)
        )
    
    if estado_filter == 'ruta':
        turnos_list = turnos_list.filter(estado='CONFIRMADO')
    elif estado_filter == 'informe':
        turnos_list = turnos_list.filter(estado='EJECUTADA')

    # 4. Paginación (9 tarjetas por página)
    paginator = Paginator(turnos_list, 9) 
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'staff/cierre.html', {
        'turnos': page_obj,
        'query': q,
        'estado_filter': estado_filter
    })

# 5. ESTADÍSTICAS
@login_required
@user_passes_test(es_staff)
def estadisticas_globales(request):
    # 1. Calcular datos inmediatamente (Server Side Rendering)
    # Esto asegura que el usuario vea los datos apenas carga la página
    stats = Turno.objects.aggregate(
        pendientes=Count('id', filter=Q(estado='PENDIENTE')),
        confirmados=Count('id', filter=Q(estado='CONFIRMADO')),
        rechazados=Count('id', filter=Q(estado='RECHAZADO')),
        terminados=Count('id', filter=Q(estado='TERMINADO')),
        cancelados=Count('id', filter=Q(estado='CANCELADO')),
        no_realizadas=Count('id', filter=Q(estado='NO_REALIZADA')),
    )
    
    # 2. Preparar datos para Chart.js
    chart_data = {
        'pie': [
            stats['terminados'], 
            stats['pendientes'], 
            stats['confirmados'], 
            stats['rechazados']
        ],
        'bar': [
            stats['terminados'], 
            stats['rechazados'], 
            stats['cancelados'], 
            stats['no_realizadas']
        ]
    }
    
    return render(request, 'staff/estadisticas.html', {
        'stats': stats,
        'chart_data_json': json.dumps(chart_data) # Enviamos JSON listo para usar
    })

@login_required
@user_passes_test(es_staff)
def gestionar_turno(request, turno_id, accion):
    with transaction.atomic():
        try:
            turno = Turno.objects.select_for_update().get(id=turno_id)
        except Turno.DoesNotExist:
            messages.error(request, "El turno no existe.")
            return redirect('dashboard_staff')

        if turno.estado != 'PENDIENTE':
            messages.warning(request, "Este turno ya fue procesado por otro inspector.")
            return redirect('dashboard_staff')

        email_usuario = turno.establecimiento.propietario.email 
        propietario = turno.establecimiento.propietario
        
        datos_email = {
            'nombre': propietario.first_name,
            'local': turno.establecimiento.nombre_comercial,
            'fecha': turno.agenda.fecha.strftime("%d/%m/%Y"),
            'jornada': turno.get_bloque_display(),
        }
        
        subject = ""
        
        if accion == 'confirmar':
            turno.estado = 'CONFIRMADO'
            turno.inspector = request.user 
            messages.success(request, f"Turno CONFIRMADO. Notificaciones enviadas.")
            
            Notificacion.objects.create(
                usuario=propietario,
                titulo="¡Turno Aprobado! ✅",
                mensaje=f"Su inspección para {turno.establecimiento.nombre_comercial} ha sido confirmada.",
                tipo="SUCCESS",
                link="/portal/"
            )
            
            subject = "Inspección Confirmada ✅"
            datos_email['mensaje_principal'] = "Nos complace informarle que su solicitud de inspección ha sido APROBADA."
            datos_email['estado'] = "CONFIRMADO"
            datos_email['color_estado'] = "#198754" # Verde
            datos_email['instrucciones'] = "Por favor, asegúrese de que una persona mayor de edad se encuentre en el establecimiento para recibir al inspector."
        
        elif accion == 'rechazar':
            turno.estado = 'RECHAZADO'
            messages.warning(request, f"Turno RECHAZADO.")
            
            Notificacion.objects.create(
                usuario=propietario,
                titulo="Solicitud Rechazada ⚠️",
                mensaje=f"No pudimos procesar su turno para {turno.establecimiento.nombre_comercial}.",
                tipo="WARNING",
                link="/portal/"
            )

            subject = "Actualización de Solicitud ⚠️"
            datos_email['mensaje_principal'] = "Le informamos que su solicitud de inspección no pudo ser procesada en la fecha seleccionada."
            datos_email['estado'] = "RECHAZADO"
            datos_email['color_estado'] = "#dc3545" # Rojo
            datos_email['instrucciones'] = "Esto puede deberse a falta de disponibilidad operativa en su zona o datos incompletos. Por favor ingrese al portal y seleccione una nueva fecha."
        
        turno.save()
    
    if email_usuario:
        enviar_correo_html(email_usuario, subject, datos_email)

    return redirect('solicitudes_pendientes')

@login_required
@user_passes_test(es_staff)
def finalizar_turno(request, turno_id):
    if request.method == 'POST':
        turno = get_object_or_404(Turno, id=turno_id)
        num = request.POST.get('numero_formulario')
        if num:
            turno.numero_formulario = num
            turno.estado = 'TERMINADO'
            turno.save()
            
            # Notificación final
            Notificacion.objects.create(
                usuario=turno.establecimiento.propietario,
                titulo="Trámite Finalizado ✅",
                mensaje=f"Proceso completado exitosamente. Formulario N° {num}.",
                tipo="SUCCESS", link="/portal/"
            )
            messages.success(request, f"Cerrado con formulario N° {num}.")
        else:
            messages.error(request, "Ingrese el número.")
    return redirect('dashboard_staff')


# ==============================================================================
#                              GENERACIÓN DE INFORMES (CORREGIDO)
# ==============================================================================

@login_required
@user_passes_test(es_staff)
def generar_informe_mensual(request):
    hoy = date.today()
    
    # 1. Obtener Parámetros
    try:
        anio = int(request.GET.get('anio', hoy.year))
        mes = int(request.GET.get('mes', hoy.month))
    except ValueError:
        anio = hoy.year
        mes = hoy.month

    tipo_reporte = request.GET.get('tipo_reporte', 'mensual') # mensual, anual, ytd
    print_mode = request.GET.get('print_mode') == '1' # ¿Es para imprimir?

    # 2. Filtro Base: Estado TERMINADO y Año seleccionado
    turnos = Turno.objects.filter(
        estado='TERMINADO',
        agenda__fecha__year=anio
    ).select_related('establecimiento__propietario', 'agenda', 'inspector').order_by('agenda__fecha')

    # 3. Aplicar Lógica de Fechas
    titulo_periodo = f"Año {anio}"
    
    if tipo_reporte == 'mensual':
        turnos = turnos.filter(agenda__fecha__month=mes)
        nombres_meses = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
        titulo_periodo = f"{nombres_meses[mes]} {anio}"
        
    elif tipo_reporte == 'ytd':
        # Hasta la fecha de hoy
        turnos = turnos.filter(agenda__fecha__lte=hoy)
        titulo_periodo = f"Enero - {hoy.strftime('%B')} {anio} (A la fecha)"
        
    elif tipo_reporte == 'anual':
        # Todo el año (ya filtrado por base)
        titulo_periodo = f"Ejercicio Fiscal {anio}"

    # Totales
    total_registros = turnos.count()

    # 4. Paginación (Solo si NO estamos en modo impresión)
    if not print_mode:
        paginator = Paginator(turnos, 20) # 20 por página en pantalla
        page_number = request.GET.get('page')
        turnos_paginados = paginator.get_page(page_number)
    else:
        turnos_paginados = turnos # Sin paginar para impresión completa

    # Contexto
    rango_anios = range(2024, hoy.year + 2)

    return render(request, 'staff/informe_mensual.html', {
        'turnos': turnos_paginados,
        'mes': mes,
        'anio': anio,
        'rango_anios': rango_anios,
        'tipo_reporte': tipo_reporte,
        'titulo_periodo': titulo_periodo,
        'total': total_registros,
        'print_mode': print_mode, # Para activar JS de impresión automática
        'anio_actual': hoy.year
    })

@login_required
@user_passes_test(es_staff)
def exportar_excel_mensual(request):
    try:
        mes = int(request.GET.get('mes', date.today().month))
        anio = int(request.GET.get('anio', date.today().year))
    except ValueError:
        mes = date.today().month
        anio = date.today().year

    # Filtro Base Excel
    turnos = Turno.objects.filter(
        estado='TERMINADO', # Ajustar a TERMINADO o FINALIZADO
        agenda__fecha__year=anio
    ).select_related('establecimiento__propietario', 'agenda', 'inspector').order_by('agenda__fecha')

    titulo_reporte = f"Inspecciones {anio}"
    if mes > 0:
        turnos = turnos.filter(agenda__fecha__month=mes)
        titulo_reporte = f"Inspecciones {mes}-{anio}"

    # Generar Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Reporte {anio}"

    font_header = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    fill_header = PatternFill(start_color='B02A37', end_color='B02A37', fill_type='solid')
    alignment_center = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.merge_cells('A1:H1')
    ws['A1'] = f"REPORTE DE INSPECCIONES - {titulo_reporte}"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = alignment_center

    headers = ['N°', 'Fecha', 'N° Formulario', 'Nombre Comercial', 'Dirección', 'RUC', 'Inspector', 'Observaciones']
    ws.append([]); ws.append(headers)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.font = font_header; cell.fill = fill_header; cell.alignment = alignment_center; cell.border = thin_border

    for idx, turno in enumerate(turnos, 1):
        inspector = f"{turno.inspector.first_name} {turno.inspector.last_name}" if turno.inspector else "--"
        # Manejo seguro de RUC (puede no tener perfil si es data antigua)
        ruc = "N/A"
        if hasattr(turno.establecimiento.propietario, 'perfil'):
             ruc = turno.establecimiento.propietario.perfil.ruc
        
        ws.append([
            idx, turno.agenda.fecha, turno.numero_formulario or "S/N",
            turno.establecimiento.nombre_comercial, turno.establecimiento.direccion,
            ruc, inspector, turno.observaciones or ""
        ])
        
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=idx+3, column=col_num)
            cell.border = thin_border
            if col_num in [1, 2, 3, 6]: cell.alignment = alignment_center

    dims = {'A': 5, 'B': 12, 'C': 15, 'D': 35, 'E': 40, 'F': 15, 'G': 25, 'H': 30}
    for col, width in dims.items(): ws.column_dimensions[col].width = width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Reporte_CBT_{titulo_reporte}.xlsx'
    wb.save(response)
    return response

@login_required
@user_passes_test(es_staff)
def hoja_ruta(request):
    origen = Point(-77.7071697, 0.8234943, srid=4326) 
    
    bloque_actual = request.GET.get('bloque', 'MANANA')
    zona_seleccionada = request.GET.get('zona', 'SUR')
    fecha_str = request.GET.get('fecha')
    
    if fecha_str:
        try:
            fecha_filtro = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        except ValueError:
            fecha_filtro = date.today()
    else:
        fecha_filtro = date.today()

    filtro_parroquia = 'TULCAN_CENTRO'
    if zona_seleccionada == 'NORTE':
        filtro_parroquia = 'GONZALEZ_SUAREZ'

    pendientes = list(Turno.objects.filter(
        agenda__fecha=fecha_filtro, 
        estado='CONFIRMADO',
        establecimiento__parroquia=filtro_parroquia,
        bloque=bloque_actual
    ).select_related('establecimiento'))

    ruta_optimizada = []
    punto_actual = origen
    
    while pendientes:
        pendientes_con_ubicacion = [p for p in pendientes if p.establecimiento.ubicacion]
        
        if not pendientes_con_ubicacion:
            break 

        siguiente = min(
            pendientes_con_ubicacion, 
            key=lambda t: t.establecimiento.ubicacion.distance(punto_actual)
        )
        ruta_optimizada.append(siguiente)
        pendientes.remove(siguiente)
        if siguiente in pendientes: pendientes.remove(siguiente) 
        punto_actual = siguiente.establecimiento.ubicacion

    return render(request, 'staff/hoja_ruta.html', {
        'ruta': ruta_optimizada,
        'hoy': fecha_filtro,
        'bloque_actual': bloque_actual,
        'zona_actual': zona_seleccionada
    })

# ==============================================================================
#                              GESTIÓN DE USUARIOS
# ==============================================================================

@login_required
@user_passes_test(es_staff)
def gestion_usuarios(request):
    query = request.GET.get('q', '')
    
    # 1. Consulta Base (Lazy Load - No ejecuta la query aún)
    if request.user.is_superuser:
        usuarios_list = User.objects.exclude(pk=request.user.pk)
    else:
        usuarios_list = User.objects.filter(is_staff=False, is_superuser=False)
    
    if query:
        usuarios_list = usuarios_list.filter(
            Q(username__icontains=query) |
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query) |
            Q(email__icontains=query) |
            Q(perfil__ruc__icontains=query)
        ).distinct()

    # Ordenar (Vital para la paginación consistente)
    usuarios_list = usuarios_list.order_by('-date_joined')

    # 2. Paginación: Mostrar 20 por página
    paginator = Paginator(usuarios_list, 20) 
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'staff/gestion_usuarios.html', {
        'page_obj': page_obj, # Enviamos el objeto paginado, no la lista completa
        'query': query
    })

@login_required
@user_passes_test(es_staff)
def detalle_usuario(request, user_id):
    usuario = get_object_or_404(User, pk=user_id)
    locales = usuario.establecimientos.all()
    return render(request, 'staff/detalle_usuario.html', {'usuario': usuario, 'locales': locales})

@login_required
@user_passes_test(es_staff)
def editar_usuario(request, user_id):
    usuario = get_object_or_404(User, pk=user_id)
    if not request.user.is_superuser and (usuario.is_staff or usuario.is_superuser):
         messages.error(request, "No tiene permisos.")
         return redirect('gestion_usuarios')

    if request.method == 'POST':
        form = EditarUsuarioForm(request.POST, instance=usuario)
        if form.is_valid():
            form.save()
            messages.success(request, "Usuario actualizado.")
            return redirect('detalle_usuario', user_id=usuario.id)
    else:
        form = EditarUsuarioForm(instance=usuario)
    return render(request, 'staff/editar_usuario.html', {'form': form, 'usuario': usuario})

@login_required
@user_passes_test(es_staff)
def eliminar_usuario(request, user_id):
    usuario = get_object_or_404(User, pk=user_id)
    if not request.user.is_superuser and usuario.is_staff:
        messages.error(request, "Acción denegada.")
        return redirect('gestion_usuarios')
    try:
        usuario.delete()
        messages.success(request, "Usuario eliminado.")
    except ProtectedError:
        messages.error(request, "No se puede eliminar: Tiene registros vinculados.")
    return redirect('gestion_usuarios')

@login_required
@user_passes_test(es_superuser)
def cambiar_rol(request, user_id):
    usuario = get_object_or_404(User, pk=user_id)
    if usuario.pk == request.user.pk:
        messages.error(request, "No puedes cambiar tu propio rol.")
        return redirect('gestion_usuarios')

    if usuario.is_staff:
        usuario.is_staff = False
        msg = f"Rol de {usuario.username} cambiado a: CIUDADANO"
    else:
        usuario.is_staff = True
        msg = f"Rol de {usuario.username} cambiado a: INSPECTOR"
    usuario.save()
    messages.success(request, msg)
    return redirect('gestion_usuarios')

@login_required
@user_passes_test(es_superuser)
def crear_inspector(request):
    if request.method == 'POST':
        form = NuevoInspectorForm(request.POST)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, "Inspector registrado correctamente.")
                return redirect('gestion_usuarios')
            except Exception as e:
                messages.error(request, f"Error al registrar: {e}")
    else:
        form = NuevoInspectorForm()
    return render(request, 'staff/crear_inspector.html', {'form': form})

# ==============================================================================
#                              OPERACIONES Y REGISTROS
# ==============================================================================

@login_required
@user_passes_test(es_staff)
def alta_contribuyente(request):
    if request.method == 'POST':
        form = AltaContribuyenteForm(request.POST)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, "Contribuyente registrado correctamente.")
                return redirect('dashboard_staff')
            except Exception as e:
                messages.error(request, f"Error al registrar: {e}")
    else:
        form = AltaContribuyenteForm()
    return render(request, 'staff/alta_contribuyente.html', {'form': form})

@login_required
def api_buscar_propietario(request):
    cedula = request.GET.get('cedula')
    response = {
        'existe': False, 
        'first_name': '', 
        'last_name': '', 
        'ruc': '', 
        'locales': []
    }
    
    if cedula and len(cedula) >= 10:
        try:
            user = User.objects.get(username=cedula)
            response['existe'] = True
            response['first_name'] = user.first_name
            response['last_name'] = user.last_name
            
            if hasattr(user, 'perfil'):
                response['ruc'] = user.perfil.ruc
            
            # Obtener locales con nombre de parroquia legible
            for loc in user.establecimientos.all():
                response['locales'].append({
                    'nombre_comercial': loc.nombre_comercial,
                    'direccion': loc.direccion,
                    'parroquia': loc.get_parroquia_display() # Nombre legible
                })
                
        except User.DoesNotExist:
            pass
            
    return JsonResponse(response)

# --- AGENDA (MASIVA) ---
@login_required
@user_passes_test(es_staff)
def habilitar_agenda(request):
    config, _ = ConfiguracionSistema.objects.get_or_create(solo_id=1)

    if request.method == 'POST':
        if 'actualizar_config' in request.POST:
            f = ConfiguracionGlobalForm(request.POST, instance=config)
            if f.is_valid(): 
                f.save()
                messages.success(request, "Configuración guardada.")
                return redirect('habilitar_agenda')
            
        elif 'crear_agenda' in request.POST:
            try:
                start = datetime.strptime(request.POST.get('fecha_inicio'), "%Y-%m-%d").date()
                end = datetime.strptime(request.POST.get('fecha_fin'), "%Y-%m-%d").date()
                zonas = request.POST.getlist('zonas')
                dias = request.POST.getlist('dias_semana')
                
                if start < date.today(): 
                    messages.error(request, "Fechas pasadas.")
                    return redirect('habilitar_agenda')

                count = 0
                notif_count = 0
                
                # Zonas Urbanas (No notificamos masivamente para no hacer spam diario)
                zonas_urbanas = ['GONZALEZ_SUAREZ', 'TULCAN_CENTRO']

                for i in range((end - start).days + 1):
                    d = start + timedelta(days=i)
                    if str(d.weekday()) in dias:
                        for z in zonas:
                            try:
                                agenda, created = AgendaDiaria.objects.get_or_create(
                                    fecha=d, 
                                    parroquia_destino=z,
                                    defaults={
                                        'capacidad_manana': config.def_capacidad_manana, 
                                        'capacidad_tarde': config.def_capacidad_tarde, 
                                        'cupos_habilitados': True
                                    }
                                )
                                
                                if created: 
                                    count += 1
                                    
                                    # --- LÓGICA DE NOTIFICACIÓN RURAL ---
                                    # Si la zona es RURAL, avisamos a todos los locales de esa zona
                                    if z not in zonas_urbanas:
                                        # Buscar usuarios con locales en esta parroquia
                                        usuarios_afectados = User.objects.filter(
                                            establecimientos__parroquia=z
                                        ).distinct()
                                        
                                        notificaciones = []
                                        fecha_fmt = d.strftime("%d/%m")
                                        zona_nombre = dict(OPCIONES_PARROQUIA).get(z, z)
                                        
                                        for u in usuarios_afectados:
                                            notificaciones.append(Notificacion(
                                                usuario=u,
                                                titulo="Visita Programada 🚒",
                                                mensaje=f"El Cuerpo de Bomberos estará en {zona_nombre} el día {fecha_fmt}. Por favor ingrese al portal y reserve su turno.",
                                                tipo="INFO",
                                                link="/portal/"
                                            ))
                                        
                                        # Guardar en bloque (Optimizado)
                                        if notificaciones:
                                            Notificacion.objects.bulk_create(notificaciones)
                                            notif_count += len(notificaciones)
                                            
                            except Exception as e:
                                print(f"Error creando agenda: {e}")
                                pass
                                
                msg = f"{count} fechas creadas."
                if notif_count > 0:
                    msg += f" Se enviaron {notif_count} alertas a ciudadanos rurales."
                
                messages.success(request, msg)
                
            except ValueError: 
                messages.error(request, "Error en formato de fechas.")
            return redirect('habilitar_agenda')

    zonas_u = [z for z in OPCIONES_PARROQUIA if z[0] in ['GONZALEZ_SUAREZ', 'TULCAN_CENTRO']]
    zonas_r = [z for z in OPCIONES_PARROQUIA if z[0] not in ['GONZALEZ_SUAREZ', 'TULCAN_CENTRO']]
    agendas = AgendaDiaria.objects.filter(fecha__gte=date.today()).order_by('fecha', 'parroquia_destino')
    
    return render(request, 'staff/habilitar_agenda.html', {
        'zonas_urbanas': zonas_u, 
        'zonas_rurales': zonas_r, 
        'agendas': agendas,
        'hoy_str': date.today().strftime("%Y-%m-%d"),
        'form_config': ConfiguracionGlobalForm(instance=config), 
        'config': config
    })

@login_required
@user_passes_test(es_staff)
def editar_agenda_detalle(request, agenda_id):
    agenda = get_object_or_404(AgendaDiaria, id=agenda_id)
    if request.method == 'POST':
        form = EdicionAgendaForm(request.POST, instance=agenda)
        if form.is_valid():
            form.save()
            messages.success(request, "Disponibilidad actualizada.")
            return redirect('habilitar_agenda')
    else:
        form = EdicionAgendaForm(instance=agenda)
    return render(request, 'staff/editar_agenda.html', {'form': form, 'agenda': agenda})

# --- TIPOS ---
@login_required
@user_passes_test(es_staff)
def gestion_tipos(request):
    if request.method == 'POST':
        form = TipoEstablecimientoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Tipo agregado correctamente.")
            return redirect('gestion_tipos')
    else:
        form = TipoEstablecimientoForm()
    tipos = TipoEstablecimiento.objects.all().order_by('nombre')
    return render(request, 'staff/gestion_tipos.html', {'form': form, 'tipos': tipos})

@login_required
@user_passes_test(es_staff)
def editar_tipo(request, tipo_id):
    tipo = get_object_or_404(TipoEstablecimiento, id=tipo_id)
    if request.method == 'POST':
        form = TipoEstablecimientoForm(request.POST, instance=tipo)
        if form.is_valid():
            form.save()
            messages.success(request, "Tipo actualizado correctamente.")
            return redirect('gestion_tipos')
    else:
        form = TipoEstablecimientoForm(instance=tipo)
    return render(request, 'staff/editar_tipo.html', {'form': form, 'tipo': tipo})

@login_required
@user_passes_test(es_staff)
def eliminar_tipo(request, tipo_id):
    tipo = get_object_or_404(TipoEstablecimiento, id=tipo_id)
    try:
        tipo.delete()
        messages.success(request, "Tipo eliminado correctamente.")
    except ProtectedError:
        messages.error(request, "⚠️ No se puede eliminar: Existen locales registrados con este tipo.")
    except Exception as e:
        messages.error(request, f"Error al eliminar: {e}")
        
    return redirect('gestion_tipos')

# --- DIRECTORIO Y DETALLES ---
@login_required
@user_passes_test(es_staff)
def directorio_establecimientos(request):
    query = request.GET.get('q', '')
    filtro_tipo = request.GET.get('tipo', '')
    
    # 1. Query Base Optimizada
    locales_list = Establecimiento.objects.all().select_related('propietario__perfil', 'tipo').order_by('nombre_comercial')
    
    # 2. Filtros
    if query:
        locales_list = locales_list.filter(
            Q(nombre_comercial__icontains=query) | 
            Q(propietario__perfil__ruc__icontains=query) | 
            Q(propietario__first_name__icontains=query)
        )
    
    if filtro_tipo:
        locales_list = locales_list.filter(tipo_id=filtro_tipo)

    # 3. Paginación (10 por página)
    paginator = Paginator(locales_list, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    tipos = TipoEstablecimiento.objects.all()
    
    return render(request, 'staff/directorio.html', {
        'locales': page_obj, # Enviamos la página actual, no toda la lista
        'tipos': tipos,
        'query': query,
        'filtro_tipo': int(filtro_tipo) if filtro_tipo else ''
    })

@login_required
@user_passes_test(es_staff)
def detalle_establecimiento(request, local_id):
    local = get_object_or_404(Establecimiento, id=local_id)
    historial = Turno.objects.filter(establecimiento=local).order_by('-agenda__fecha')
    
    if request.method == 'POST':
        local.nombre_comercial = request.POST.get('nombre_comercial').upper()
        local.razon_social = request.POST.get('razon_social').upper()
        local.direccion = request.POST.get('direccion').upper()
        local.parroquia = request.POST.get('parroquia')
        
        tipo_id = request.POST.get('tipo')
        if tipo_id:
            local.tipo_id = tipo_id

        lat = request.POST.get('latitud')
        lon = request.POST.get('longitud')
        if lat and lon:
            try:
                local.ubicacion = Point(float(lon), float(lat), srid=4326)
                local.ubicacion_verificada = True 
            except ValueError:
                pass

        local.save()
        messages.success(request, "Ficha del establecimiento actualizada correctamente.")
        return redirect('detalle_establecimiento', local_id=local.id)

    tipos = TipoEstablecimiento.objects.all().order_by('nombre')
    return render(request, 'staff/detalle_local.html', {
        'local': local, 
        'historial': historial,
        'tipos': tipos,
        'parroquias': OPCIONES_PARROQUIA
    })

# --- VENTANILLA Y AGENDAMIENTO ---
@login_required
@user_passes_test(es_staff)
def buscar_local_presencial(request):
    query = request.GET.get('q')
    locales = []
    if query:
        locales = Establecimiento.objects.filter(
            Q(propietario__perfil__ruc__icontains=query) |
            Q(nombre_comercial__icontains=query) |
            Q(razon_social__icontains=query)
        )[:10]

    return render(request, 'staff/buscar_local.html', {'locales': locales, 'query': query})

@login_required
@user_passes_test(es_staff)
def agendar_presencial_detalle(request, local_id):
    local = get_object_or_404(Establecimiento, id=local_id)
    
    # 1. VALIDACIÓN: ¿Tiene turno activo?
    turno_activo = Turno.objects.filter(
        establecimiento=local,
        estado__in=['PENDIENTE', 'CONFIRMADO'],
        agenda__fecha__gte=date.today()
    ).first()
    
    if turno_activo:
        return render(request, 'staff/agendar_presencial.html', {
            'local': local,
            'turno_activo': turno_activo,
            'opciones': None
        })

    # 2. OBTENER AGENDA (Ordenada por fecha)
    agendas = AgendaDiaria.objects.filter(
        parroquia_destino=local.parroquia,
        fecha__gte=date.today(),
        cupos_habilitados=True
    ).order_by('fecha')

    # 3. PROCESAR GUARDADO
    if request.method == 'POST':
        with transaction.atomic():
            agenda = AgendaDiaria.objects.select_for_update().get(id=request.POST.get('agenda_id'))
            bloque = request.POST.get('bloque')
            
            # Validar Cupo Real
            ocupados = Turno.objects.filter(agenda=agenda, bloque=bloque).exclude(estado='CANCELADO').count()
            capacidad = agenda.capacidad_manana if bloque == 'MANANA' else agenda.capacidad_tarde
            
            if ocupados >= capacidad:
                messages.error(request, "❌ Error: El cupo seleccionado acaba de llenarse.")
                return redirect('agendar_presencial_detalle', local_id=local.id)

            Turno.objects.create(
                agenda=agenda,
                establecimiento=local,
                bloque=bloque,
                estado='CONFIRMADO',
                inspector=request.user,
                observaciones="VENTANILLA",
                telefono_contacto=request.POST.get('telefono'),
                referencia_ubicacion=request.POST.get('referencia')
            )
            
            Notificacion.objects.create(
                usuario=local.propietario,
                titulo="Turno Asignado ✅",
                mensaje=f"Confirmado turno presencial para {local.nombre_comercial}.",
                tipo="SUCCESS", link="/portal/"
            )
            messages.success(request, "Turno confirmado exitosamente.")
            return redirect('dashboard_staff')

    # 4. CALCULAR CUPOS
    opciones = []
    for ag in agendas:
        ocup_man = Turno.objects.filter(agenda=ag, bloque='MANANA').exclude(estado='CANCELADO').count()
        ocup_tar = Turno.objects.filter(agenda=ag, bloque='TARDE').exclude(estado='CANCELADO').count()
        bloques = []
        
        if ag.capacidad_manana > 0:
            pct = int((ocup_man / ag.capacidad_manana) * 100)
            bloques.append({
                'codigo': 'MANANA', 'label': 'MAÑANA', 'hora': '09:00 - 12:30', 
                'ocupados': ocup_man, 'total': ag.capacidad_manana, 'pct': pct, 
                'disponible': ocup_man < ag.capacidad_manana
            })
            
        if ag.capacidad_tarde > 0:
            pct = int((ocup_tar / ag.capacidad_tarde) * 100)
            bloques.append({
                'codigo': 'TARDE', 'label': 'TARDE', 'hora': '14:45 - 16:30', 
                'ocupados': ocup_tar, 'total': ag.capacidad_tarde, 'pct': pct, 
                'disponible': ocup_tar < ag.capacidad_tarde
            })
            
        if bloques:
            opciones.append({'info': ag, 'bloques': bloques})

    return render(request, 'staff/agendar_presencial.html', {'local': local, 'opciones': opciones})

# ==============================================================================
#                              PORTAL CIUDADANO
# ==============================================================================

@login_required
def api_mis_notificaciones(request):
    notifs = Notificacion.objects.filter(usuario=request.user, leido=False)[:5]
    data = [{'id': n.id, 'titulo': n.titulo, 'mensaje': n.mensaje, 'tipo': n.tipo, 'fecha': n.fecha_creacion.strftime("%H:%M"), 'link': n.link} for n in notifs]
    return JsonResponse({'count': len(data), 'notificaciones': data})

@login_required
def api_marcar_leida(request, notificacion_id):
    try:
        n = Notificacion.objects.get(id=notificacion_id, usuario=request.user)
        n.leido = True
        n.save()
        return JsonResponse({'status': 'ok'})
    except Notificacion.DoesNotExist:
        return JsonResponse({'status': 'error'}, status=404)

@login_required
def home_ciudadano(request):
    if request.user.is_staff:
        return redirect('dashboard_staff')

    # ... (Validaciones de email y mapa se mantienen igual) ...
    if not request.user.email: return redirect('registrar_email')
    mis_locales = request.user.establecimientos.all()
    for local in mis_locales:
        if not local.ubicacion_verificada: return redirect('verificar_ubicacion', local_id=local.id)

    # LÓGICA ROBUSTA DE FILTRADO
    today = date.today()
    
    selected_id = request.GET.get('local_id')
    selected_local = mis_locales.filter(id=selected_id).first() if selected_id else mis_locales.first()
    
    # 1. EN CURSO: Solo turnos FUTUROS o de HOY que sean PENDIENTE/CONFIRMADO
    # Esto previene que un turno de ayer aparezca aquí aunque el script de limpieza no haya corrido aún.
    turnos_activos = Turno.objects.filter(
        establecimiento__in=mis_locales,
        agenda__fecha__gte=today, # <--- FILTRO CRÍTICO: Mayor o igual a hoy
        estado__in=['PENDIENTE', 'CONFIRMADO']
    ).order_by('agenda__fecha')
    
    # 2. HISTORIAL: Todo lo demás
    # Incluye: Pasados, Rechazados, Terminados, Cancelados, No Realizados
    historial = Turno.objects.filter(
        establecimiento__in=mis_locales
    ).exclude(
        id__in=turnos_activos
    ).order_by('-agenda__fecha')

    return render(request, 'ciudadano/home.html', {
        'locales': mis_locales,
        'selected_local': selected_local,
        'turnos_activos': turnos_activos,
        'historial': historial
    })

@login_required
def registrar_email(request):
    if request.user.email: return redirect('home_ciudadano')
    if request.method == 'POST':
        form = RegistroEmailForm(request.POST, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, "Correo registrado.")
            return redirect('home_ciudadano')
    else:
        form = RegistroEmailForm(instance=request.user)
    return render(request, 'ciudadano/registrar_email.html', {'form': form})

@login_required
def verificar_ubicacion(request, local_id):
    local = get_object_or_404(Establecimiento, id=local_id, propietario=request.user)
    if request.method == 'POST':
        if request.POST.get('latitud'):
            local.ubicacion = Point(float(request.POST.get('longitud')), float(request.POST.get('latitud')), srid=4326)
            local.ubicacion_verificada = True; local.save()
            messages.success(request, "Ubicación guardada.")
            return redirect('home_ciudadano')
    return render(request, 'ciudadano/verificar_ubicacion.html', {'local': local})

@login_required
def detalle_local_ciudadano(request, local_id):
    local = get_object_or_404(Establecimiento, id=local_id, propietario=request.user)
    historial = Turno.objects.filter(establecimiento=local).order_by('-agenda__fecha')
    return render(request, 'ciudadano/detalle_local.html', {'local': local, 'historial': historial})

@login_required
def agendar_turno(request):
    local = get_object_or_404(Establecimiento, id=request.GET.get('local_id'), propietario=request.user)
    
    # Validación Ciudadano
    turno_activo = Turno.objects.filter(
        establecimiento=local,
        estado__in=['PENDIENTE', 'CONFIRMADO'],
        agenda__fecha__gte=date.today()
    ).first()
    
    if turno_activo:
        messages.warning(request, f"Este local ya tiene una solicitud activa.")
        return render(request, 'ciudadano/agendar.html', {'local': local, 'turno_activo': turno_activo, 'opciones': None})

    agendas = AgendaDiaria.objects.filter(
        parroquia_destino=local.parroquia,
        fecha__gte=date.today(),
        cupos_habilitados=True
    ).order_by('fecha')

    if request.method == 'POST':
        with transaction.atomic():
            agenda = AgendaDiaria.objects.select_for_update().get(id=request.POST.get('agenda_id'))
            bloque = request.POST.get('bloque')
            
            ocupados = Turno.objects.filter(agenda=agenda, bloque=bloque).exclude(estado='CANCELADO').count()
            cap = agenda.capacidad_manana if bloque == 'MANANA' else agenda.capacidad_tarde
            
            if ocupados >= cap:
                messages.error(request, "Cupo lleno.")
                return redirect(f"/portal/agendar/?local_id={local.id}")
            
            Turno.objects.create(
                agenda=agenda, establecimiento=local, bloque=bloque,
                telefono_contacto=request.POST.get('telefono'),
                referencia_ubicacion=request.POST.get('referencia'),
                estado='PENDIENTE'
            )
            
            # Notificar Staff
            for insp in User.objects.filter(is_staff=True):
                Notificacion.objects.create(
                    usuario=insp,
                    titulo="Nueva Solicitud 📥",
                    mensaje=f"{local.nombre_comercial} ha solicitado turno.",
                    tipo="INFO",
                    link="/panel-operativo/"
                )

            messages.success(request, "Solicitud enviada.")
            return redirect('home_ciudadano')

    opciones = []
    for ag in agendas:
        ocup_man = Turno.objects.filter(agenda=ag, bloque='MANANA').exclude(estado='CANCELADO').count()
        ocup_tar = Turno.objects.filter(agenda=ag, bloque='TARDE').exclude(estado='CANCELADO').count()
        
        bloques = []
        if ag.capacidad_manana > 0:
             pct = int((ocup_man/ag.capacidad_manana)*100)
             bloques.append({'codigo': 'MANANA', 'label': 'MAÑANA', 'hora': '09:00 - 12:30', 'ocupados': ocup_man, 'total': ag.capacidad_manana, 'pct': pct, 'disponible': ocup_man < ag.capacidad_manana})
        if ag.capacidad_tarde > 0:
             pct = int((ocup_tar/ag.capacidad_tarde)*100)
             bloques.append({'codigo': 'TARDE', 'label': 'TARDE', 'hora': '14:45 - 16:30', 'ocupados': ocup_tar, 'total': ag.capacidad_tarde, 'pct': pct, 'disponible': ocup_tar < ag.capacidad_tarde})
        if bloques: opciones.append({'info': ag, 'bloques': bloques})

    return render(request, 'ciudadano/agendar.html', {'local': local, 'opciones': opciones})

@login_required
def cancelar_turno(request, turno_id):
    turno = get_object_or_404(Turno, id=turno_id)
    
    if not request.user.is_staff and turno.establecimiento.propietario != request.user:
        messages.error(request, "No tiene permiso.")
        return redirect('home_ciudadano')

    if turno.agenda.fecha <= date.today():
        messages.error(request, "No se puede cancelar el mismo día.")
        if request.user.is_staff: return redirect('dashboard_staff')
        return redirect('home_ciudadano')

    turno.estado = 'CANCELADO'
    turno.save()
    
    if request.user.is_staff:
        Notificacion.objects.create(
            usuario=turno.establecimiento.propietario,
            titulo="Turno Cancelado",
            mensaje=f"Su turno para {turno.establecimiento.nombre_comercial} ha sido cancelado.",
            tipo="WARNING"
        )
    
    messages.success(request, "El turno ha sido cancelado.")
    
    if request.user.is_staff: return redirect('dashboard_staff')
    return redirect('home_ciudadano')

@login_required
def mi_perfil(request):
    user = request.user
    perfil = getattr(user, 'perfil', None)
    
    dias = 0
    puede = True
    
    if perfil and perfil.fecha_ultima_actualizacion:
        f_unlock = perfil.fecha_ultima_actualizacion + timedelta(days=90)
        if date.today() < f_unlock:
            puede = False
            dias = (f_unlock - date.today()).days
    
    if request.method == 'POST':
        if not puede and not user.is_superuser:
            messages.error(request, f"Edición bloqueada. Espere {dias} días.")
            return redirect('mi_perfil')

        form = MiPerfilForm(request.POST, instance=user)
        if form.is_valid():
            form.save()
            messages.success(request, "Perfil actualizado.")
            return redirect('mi_perfil')
    else:
        form = MiPerfilForm(instance=user)

    extra = {'puede_editar': puede, 'dias_restantes': dias}
    if user.is_staff:
        extra['total_inspecciones'] = Turno.objects.filter(inspector=user, estado='CONFIRMADO').count()
        extra['tipo_usuario'] = 'INSPECTOR'
    else:
        extra['locales'] = user.establecimientos.all()
        extra['tipo_usuario'] = 'CIUDADANO'

    return render(request, 'perfil.html', {'form': form, **extra})

@login_required
def ver_tasas_impuestos(request):
    return render(request, 'ciudadano/docs/tasas.html', {'anio_fiscal': date.today().year, 'tasas': TasaPago.objects.all()})

@login_required
def ver_guia_requisitos(request):
    return render(request, 'ciudadano/docs/requisitos.html', {'anio_fiscal': date.today().year, 'requisitos': RequisitoLegal.objects.all()})

@login_required
def api_mis_notificaciones(request):
    notifs = Notificacion.objects.filter(usuario=request.user, leido=False)[:5]
    data = [{'id': n.id, 'titulo': n.titulo, 'mensaje': n.mensaje, 'tipo': n.tipo, 'link': n.link, 'fecha': n.fecha_creacion.strftime("%H:%M")} for n in notifs]
    return JsonResponse({'count': len(data), 'notificaciones': data})

@login_required
def api_marcar_leida(request, notificacion_id):
    try: n = Notificacion.objects.get(id=notificacion_id, usuario=request.user); n.leido = True; n.save(); return JsonResponse({'status': 'ok'})
    except: return JsonResponse({'status': 'error'}, status=404)

@login_required
@user_passes_test(es_staff)
def gestion_documentacion(request):
    # 1. Listas Completas (Para Impresión)
    all_tasas = TasaPago.objects.all()
    all_requisitos = RequisitoLegal.objects.all()
    
    # 2. Paginación TASAS (10 por página)
    paginator_tasas = Paginator(all_tasas, 10)
    page_tasa_num = request.GET.get('page_tasa')
    tasas_page = paginator_tasas.get_page(page_tasa_num)

    # 3. Paginación REQUISITOS (10 por página)
    # Nota: Regroup en template funcionará sobre la página actual, lo cual es correcto para gestión.
    paginator_req = Paginator(all_requisitos, 10)
    page_req_num = request.GET.get('page_req')
    requisitos_page = paginator_req.get_page(page_req_num)
    
    form_tasa = TasaPagoForm()
    form_req = RequisitoLegalForm()

    if request.method == 'POST':
        if 'crear_tasa' in request.POST:
            f = TasaPagoForm(request.POST)
            if f.is_valid(): 
                f.save()
                messages.success(request, "Tasa agregada.")
                return redirect('gestion_documentacion')
            
        elif 'crear_requisito' in request.POST:
            f = RequisitoLegalForm(request.POST)
            if f.is_valid(): 
                f.save()
                messages.success(request, "Requisito agregado.")
                return redirect('gestion_documentacion')

    return render(request, 'staff/gestion_documentacion.html', {
        'tasas': tasas_page,             # Paginado (Vista)
        'requisitos': requisitos_page,   # Paginado (Vista)
        'all_tasas': all_tasas,          # Completo (Impresión)
        'all_requisitos': all_requisitos,# Completo (Impresión)
        'form_tasa': form_tasa,
        'form_req': form_req
    })

@login_required
@user_passes_test(es_staff)
def eliminar_documento(request, tipo, id_obj):
    try:
        if tipo == 'tasa': TasaPago.objects.get(id=id_obj).delete()
        elif tipo == 'requisito': RequisitoLegal.objects.get(id=id_obj).delete()
        messages.success(request, "Eliminado.")
    except: messages.error(request, "Error.")
    return redirect('gestion_documentacion')

@login_required
def api_estadisticas(request):
    """
    Devuelve estadísticas con caché de 30 segundos para evitar saturación.
    """
    # Intentar obtener de caché
    cache_key = 'dashboard_stats_global'
    stats = cache.get(cache_key)

    if not stats:
        # Si no existe, calcular
        stats = Turno.objects.aggregate(
            pendientes=Count('id', filter=Q(estado='PENDIENTE')),
            confirmados=Count('id', filter=Q(estado='CONFIRMADO')),
            rechazados=Count('id', filter=Q(estado='RECHAZADO')),
            terminados=Count('id', filter=Q(estado='TERMINADO')),
            cancelados=Count('id', filter=Q(estado='CANCELADO')),
            no_realizadas=Count('id', filter=Q(estado='NO_REALIZADA')),
        )
        # Guardar en caché por 30 segundos
        cache.set(cache_key, stats, 30)
    
    return JsonResponse(stats)

@login_required
@user_passes_test(es_staff)
def reportar_ausencia(request, turno_id):
    if request.method == 'POST':
        turno = get_object_or_404(Turno, id=turno_id)
        
        # Solo se puede reportar ausencia el MISMO DÍA de la agenda
        if turno.agenda.fecha != date.today():
            messages.error(request, "Solo puede reportar ausencia el día de la inspección.")
            return redirect('dashboard_staff')

        turno.estado = 'AUSENTE'
        turno.observaciones = "El inspector acudió al sitio pero no hubo atención."
        turno.save()
        
        # Notificar al ciudadano
        Notificacion.objects.create(
            usuario=turno.establecimiento.propietario,
            titulo="Visita Fallida 🏠",
            mensaje="Nuestro inspector visitó su local hoy pero no fue atendido.",
            tipo="WARNING",
            link="/portal/"
        )
        
        messages.warning(request, f"Turno marcado como CLIENTE AUSENTE.")
            
    return redirect('dashboard_staff')

def error_404(request, exception):
    """Muestra la página de página no encontrada."""
    return render(request, '404.html', status=404)

def error_500(request):
    """Muestra la página de error interno del servidor."""
    return render(request, '500.html', status=500)